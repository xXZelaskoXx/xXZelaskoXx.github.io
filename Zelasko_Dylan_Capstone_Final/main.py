"""
file name: main.py
purpose: call applicable functions to fulfill parameters of final project for CITC Capstone
name: Dylan Zelasko
"""
from tkinter import Button

#Used to read and write storage file
import openpyxl
import tkinter as tk

#Constants for project will be stored in the A line of spreadsheet document

#Containers for A and B elements
AContainer = list()
BContainer = list()

#Open container worksheet for RadioLog
wb = openpyxl.load_workbook('sheets/radioLog.xlsx')

#set active worksheet as current scope
ws = wb.active

#BUILD GUI <tk>

#Establish main window container
window = tk.Tk()
window.configure(bg="#B33B72")
window.title("RAD-ioLog: Only the RADDEST inputs and outputs")

"""
function: saveFile()
purpose: sets all values of cells to most recent values in GUI, then saves workbook
params: none
return: none
"""
def saveFile():

#Set Values for row A
    ws['A2'] = A2Entry.get()
    ws['A5'] = A5Entry.get()
    ws['A7'] = A7Entry.get()
    ws['A9'] = A9Entry.get()

#Set Values for Row C
    for x in range(12):
        cellNumber = "C" + str(x + 1)
        ws[cellNumber] = CContainer[x].get()
#Set Values for Row D
        cellNumber = "D" + str(x + 1)
        ws[cellNumber] = DContainer[x].get()
#Set Values for Row E
        cellNumber = "E" + str(x + 1)
        ws[cellNumber] = EContainer[x].get()
#Set Values for Row F
        cellNumber = "F" + str(x + 1)
        ws[cellNumber] = FContainer[x].get()
#Set Values for Row G
        cellNumber = "G" + str(x + 1)
        ws[cellNumber] = GContainer[x].get()
#Set Values for Row H
        cellNumber = "H" + str(x + 1)
        ws[cellNumber] = HContainer[x].get()
#Set Values for Row I
        cellNumber = "I" + str(x + 1)
        ws[cellNumber] = IContainer[x].get()
#Set Values for Row J
        cellNumber = "J" + str(x + 1)
        ws[cellNumber] = JContainer[x].get()
#Set Values for Row K
        cellNumber = "K" + str(x + 1)
        ws[cellNumber] = KContainer[x].get()
#Set Values for Row L
        cellNumber = "L" + str(x + 1)
        ws[cellNumber] = LContainer[x].get()
#Set Values for Row M
        cellNumber = "M" + str(x + 1)
        ws[cellNumber] = MContainer[x].get()
#Set Values for Row N
        cellNumber = "N" + str(x + 1)
        ws[cellNumber] = NContainer[x].get()
#Set Values for Row O:
        cellNumber = "O" + str(x + 1)
        ws[cellNumber] = OContainer[x].get()
#Set Values for Row P
        cellNumber = "P" + str(x + 1)
        ws[cellNumber] = PContainer[x].get()
#Set Values for Row Q
        cellNumber = "Q" + str(x + 1)
        ws[cellNumber] = QContainer[x].get()

#SAVE file
    wb.save('sheets/radioLog.xlsx')

#END saveFile()

"""
function: lightMode()
purpose: changes style rule to light mode for GUI
params: none
return: none
"""
def lightMode():

#COLORS: window bg: #B33B72, fg="#4969E1", bg="pink"
    window.configure(bg="#B33B72")

#apply light mode style to button elements
    LightModeButton.configure(fg="#4969E1", bg="pink")
    DarkModeButton.configure(fg="#4969E1", bg="pink")
    SaveButton.configure(fg="#4969E1", bg="pink")

#apply light mode style to search elements
    searchLabel.configure(fg="#4969E1", bg="pink")
    searchEntry.configure(fg="#4969E1", bg="pink")
    searchButton.configure(fg="#4969E1", bg="pink")

#for loop to iterate all elements within containers for light mode style
    for i in range(12):
        if(i<9):
            AContainer[i].configure(fg="#4969E1", bg="pink")
        BContainer[i].configure(fg="#4969E1", bg="pink")
        CContainer[i].configure(fg="#4969E1", bg="pink")
        DContainer[i].configure(fg="#4969E1", bg="pink")
        EContainer[i].configure(fg="#4969E1", bg="pink")
        FContainer[i].configure(fg="#4969E1", bg="pink")
        GContainer[i].configure(fg="#4969E1", bg="pink")
        HContainer[i].configure(fg="#4969E1", bg="pink")
        IContainer[i].configure(fg="#4969E1", bg="pink")
        JContainer[i].configure(fg="#4969E1", bg="pink")
        KContainer[i].configure(fg="#4969E1", bg="pink")
        LContainer[i].configure(fg="#4969E1", bg="pink")
        MContainer[i].configure(fg="#4969E1", bg="pink")
        NContainer[i].configure(fg="#4969E1", bg="pink")
        OContainer[i].configure(fg="#4969E1", bg="pink")
        PContainer[i].configure(fg="#4969E1", bg="pink")
        QContainer[i].configure(fg="#4969E1", bg="pink")
#END lightMode()



"""
function: darkMode()
purpose: changes style rule to dark mode for GUI
params: none
return: none
"""
def darkMode():
#dark mode colors bg: black, fg: #87CEEB, bg: #363737

#set main window to bg: black
    window.configure(bg="black")

#apply dark mode style to button elements
    LightModeButton.configure(fg="#87CEEB", bg="#363737")
    DarkModeButton.configure(fg="#87CEEB", bg="#363737")
    SaveButton.configure(fg="#87CEEB", bg="#363737")

#apply dark mode style to search elements
    searchLabel.configure(fg="#87CEEB", bg="#363737")
    searchEntry.configure(fg="#87CEEB", bg="#363737")
    searchButton.configure(fg="#87CEEB", bg="#363737")

#iterate through all container classes and apply dark mode style
    for i in range(12):

        if (i < 9):
            AContainer[i].configure(fg="#87CEEB", bg="#363737")
        BContainer[i].configure(fg="#87CEEB", bg="#363737")
        CContainer[i].configure(fg="#87CEEB", bg="#363737")
        DContainer[i].configure(fg="#87CEEB", bg="#363737")
        EContainer[i].configure(fg="#87CEEB", bg="#363737")
        FContainer[i].configure(fg="#87CEEB", bg="#363737")
        GContainer[i].configure(fg="#87CEEB", bg="#363737")
        HContainer[i].configure(fg="#87CEEB", bg="#363737")
        IContainer[i].configure(fg="#87CEEB", bg="#363737")
        JContainer[i].configure(fg="#87CEEB", bg="#363737")
        KContainer[i].configure(fg="#87CEEB", bg="#363737")
        LContainer[i].configure(fg="#87CEEB", bg="#363737")
        MContainer[i].configure(fg="#87CEEB", bg="#363737")
        NContainer[i].configure(fg="#87CEEB", bg="#363737")
        OContainer[i].configure(fg="#87CEEB", bg="#363737")
        PContainer[i].configure(fg="#87CEEB", bg="#363737")
        QContainer[i].configure(fg="#87CEEB", bg="#363737")
#END darkMode()


"""
function: searchFile()
purpose: takes inputs for name, call-sign, and date, then compares them against relevant values
    in fields. If there is a match, the style is updated to show such.
params: none
return: none
"""
def searchFile():

#For loop to iterate through every searchable element
    for i in range(12):
    #check C values
       if searchEntry.get().lower() in CContainer[i].get().lower():
           CContainer[i].configure(fg="black" ,bg="#FFFF00")
    #check D values
       if searchEntry.get().lower() in DContainer[i].get().lower():
           DContainer[i].configure(fg="black" ,bg="#FFFF00")
    #check E values
       if searchEntry.get().lower() in EContainer[i].get().lower():
           EContainer[i].configure(fg="black" ,bg="#FFFF00")
    #check F values
       if searchEntry.get().lower() in FContainer[i].get().lower():
           FContainer[i].configure(fg="black" ,bg="#FFFF00")
    #check G values
       if searchEntry.get().lower() in GContainer[i].get().lower():
           GContainer[i].configure(fg="black" ,bg="#FFFF00")
    #check H values
       if searchEntry.get().lower() in HContainer[i].get().lower():
           HContainer[i].configure(fg="black" ,bg="#FFFF00")
    #check I values
       if searchEntry.get().lower() in IContainer[i].get().lower():
           IContainer[i].configure(fg="black" ,bg="#FFFF00")
    #check J values
       if searchEntry.get().lower() in JContainer[i].get().lower():
           JContainer[i].configure(fg="black" ,bg="#FFFF00")
    #check K values
       if searchEntry.get().lower() in KContainer[i].get().lower():
           KContainer[i].configure(fg="black" ,bg="#FFFF00")
    #check L values
       if searchEntry.get().lower() in LContainer[i].get().lower():
           LContainer[i].configure(fg="black" ,bg="#FFFF00")
    #check M values
       if searchEntry.get().lower() in MContainer[i].get().lower():
           MContainer[i].configure(fg="black" ,bg="#FFFF00")
    #check N values
       if searchEntry.get().lower() in NContainer[i].get().lower():
           NContainer[i].configure(fg="black" ,bg="#FFFF00")
    #check O values
       if searchEntry.get().lower() in OContainer[i].get().lower():
           OContainer[i].configure(fg="black" ,bg="#FFFF00")
    #check P values
       if searchEntry.get().lower() in PContainer[i].get().lower():
           PContainer[i].configure(fg="black" ,bg="#FFFF00")
    #check Q values
       if searchEntry.get().lower() in QContainer[i].get().lower():
           QContainer[i].configure(fg="black" ,bg="#FFFF00")
#end searchFile()


"""
function: main
purpose: constructs GUI components and interfaces with previously defined methods
    to update and use the data sheet as needed per use case.
"""
#Label for Equipment Line at A1
A1Label = tk.Label(window,
                     text=ws['A1'].value,
                     justify="center",
                     fg="#4969E1",
                     bg="pink"
)
A1Label.grid(row=0, column=0, pady=20)
AContainer.append(A1Label)

#Entry for Equipment Entry at A2
A2Entry = tk.Entry(window,
                   fg="#4969E1",
                   bg="pink")
A2Entry.insert(0, ws['A2'].value)
A2Entry.grid(row=0, column=1, sticky='W')
AContainer.append(A2Entry)

#Label for A3 Centered Element
A3Label = tk.Label(window,
                   text=ws['A3'].value,
                   justify="center",
                   fg="#4969E1",
                   bg="pink")
A3Label.grid(row=0, column=2)
AContainer.append(A3Label)

#Label for A4 "Grid: "
A4Label = tk.Label(window,
                   text=ws['A4'].value,
                   justify="center",
                   fg="#4969E1",
                   bg="pink")
A4Label.grid(row=0, column=3)
AContainer.append(A4Label)

#Entry for Grid Entry at A5
A5Entry = tk.Entry(window,
                   fg="#4969E1",
                   bg="pink")
A5Entry.insert(0, ws['A5'].value)
A5Entry.grid(row=0, column=4)
AContainer.append(A5Entry)

#Label for A6 "CQ Zone: "
A6Label = tk.Label(window,
                   text=ws['A6'].value,
                   justify="center",
                   fg="#4969E1",
                   bg="pink")
A6Label.grid(row=0, column=5)
AContainer.append(A6Label)

#Entry for CQ Zone Entry at A7
A7Entry = tk.Entry(window,
                   fg="#4969E1",
                   bg="pink")
A7Entry.insert(0, ws['A7'].value)
A7Entry.grid(row=0, column=6)
AContainer.append(A7Entry)

#Label for A8 "ITU Zone: "
A8Label = tk.Label(window,
                   text=ws['A8'].value,
                   justify="center",
                   fg="#4969E1",
                   bg="pink")
A8Label.grid(row=0, column=7)
AContainer.append(A8Label)

#Entry for ITU Zone Entry at A7
A9Entry = tk.Entry(window,
                   fg="#4969E1",
                   bg="pink")
A9Entry.insert(0, ws['A9'].value)
A9Entry.grid(row=0, column=8)
AContainer.append(A9Entry)

#Initialize B1-B12 from ws for Headers

#Label for B1" "
B1Label = tk.Label(window,
                   text=ws['B1'].value,
                   justify="center",
                   fg="#4969E1",
                   bg="pink")
B1Label.grid(row=1, column=0)
BContainer.append(B1Label)

#Label for B " "
B2Label = tk.Label(window,
                   text=ws['B2'].value,
                   justify="center",
                   fg="#4969E1",
                   bg="pink")
B2Label.grid(row=1, column=1, pady=5)
BContainer.append(B2Label)

#Label for B " "
B3Label = tk.Label(window,
                   text=ws['B3'].value,
                   justify="center",
                   fg="#4969E1",
                   bg="pink")
B3Label.grid(row=1, column=2)
BContainer.append(B3Label)

#Label for B " "
B4Label = tk.Label(window,
                   text=ws['B4'].value,
                   justify="center",
                   fg="#4969E1",
                   bg="pink")
B4Label.grid(row=1, column=3)
BContainer.append(B4Label)

#Label for B " "
B5Label = tk.Label(window,
                   text=ws['B5'].value,
                   justify="center",
                   fg="#4969E1",
                   bg="pink")
B5Label.grid(row=1, column=4)
BContainer.append(B5Label)

#Label for B " "
B6Label = tk.Label(window,
                   text=ws['B6'].value,
                   justify="center",
                   fg="#4969E1",
                   bg="pink")
B6Label.grid(row=1, column=5)
BContainer.append(B6Label)

#Label for B " "
B7Label = tk.Label(window,
                   text=ws['B7'].value,
                   justify="center",
                   fg="#4969E1",
                   bg="pink")
B7Label.grid(row=1, column=6)
BContainer.append(B7Label)

#Label for B " "
B8Label = tk.Label(window,
                   text=ws['B8'].value,
                   justify="center",
                   fg="#4969E1",
                   bg="pink")
B8Label.grid(row=1, column=7)
BContainer.append(B8Label)

#Label for B " "
B9Label = tk.Label(window,
                   text=ws['B9'].value,
                   justify="center",
                   fg="#4969E1",
                   bg="pink")
B9Label.grid(row=1, column=8)
BContainer.append(B9Label)

#Label for B " "
B10Label = tk.Label(window,
                   text=ws['B10'].value,
                   justify="center",
                   fg="#4969E1",
                   bg="pink")
B10Label.grid(row=1, column=9)
BContainer.append(B10Label)

#Label for B " "
B11Label = tk.Label(window,
                   text=ws['B11'].value,
                   justify="center",
                   fg="#4969E1",
                   bg="pink")
B11Label.grid(row=1, column=10)
BContainer.append(B11Label)

#Label for B " "
B12Label = tk.Label(window,
                   text=ws['B12'].value,
                   justify="center",
                   fg="#4969E1",
                   bg="pink")
B12Label.grid(row=1, column=11)
BContainer.append(B12Label)


#Containers for all elements C through Q
CContainer = list()
DContainer = list()
EContainer = list()
FContainer = list()
GContainer = list()
HContainer = list()
IContainer = list()
JContainer = list()
KContainer = list()
LContainer = list()
MContainer = list()
NContainer = list()
OContainer = list()
PContainer = list()
QContainer = list()

#For loop to iterate through every element C through Q
for x in range(12):

    # C1 through C12
    CEntry = tk.Entry(window,
                   fg="#4969E1",
                   bg="pink")
    CEntry.insert(0, ws['C' + str(x+1)].value)
    CContainer.append(CEntry)
    CContainer[x].grid(row=2, column=x)

#D1 through D12
    DEntry = tk.Entry(window,
                   fg="#4969E1",
                   bg="pink")
    DEntry.insert(0, ws['D' + str(x+1)].value)
    DContainer.append(DEntry)
    DContainer[x].grid(row=3, column=x)

#E1 through E12
    EEntry = tk.Entry(window,
                   fg="#4969E1",
                   bg="pink")
    EEntry.insert(0, ws['E' + str(x+1)].value)
    EContainer.append(EEntry)
    EContainer[x].grid(row=4, column=x)

#F1 through F12
    FEntry = tk.Entry(window,
                   fg="#4969E1",
                   bg="pink")
    FEntry.insert(0, ws['F' + str(x+1)].value)
    FContainer.append(FEntry)
    FContainer[x].grid(row=5, column=x)

#G1 through G12
    GEntry = tk.Entry(window,
                   fg="#4969E1",
                   bg="pink")
    GEntry.insert(0, ws['G' + str(x+1)].value)
    GContainer.append(GEntry)
    GContainer[x].grid(row=6, column=x)

#H1 through H12
    HEntry = tk.Entry(window,
                   fg="#4969E1",
                   bg="pink")
    HEntry.insert(0, ws['H' + str(x+1)].value)
    HContainer.append(HEntry)
    HContainer[x].grid(row=7, column=x)

#I1 through I12
    IEntry = tk.Entry(window,
                   fg="#4969E1",
                   bg="pink")
    IEntry.insert(0, ws['I' + str(x+1)].value)
    IContainer.append(IEntry)
    IContainer[x].grid(row=8, column=x)

#J1 through J12
    JEntry = tk.Entry(window,
                   fg="#4969E1",
                   bg="pink")
    JEntry.insert(0, ws['J' + str(x+1)].value)
    JContainer.append(JEntry)
    JContainer[x].grid(row=9, column=x)

#K1 through K12
    KEntry = tk.Entry(window,
                   fg="#4969E1",
                   bg="pink")
    KEntry.insert(0, ws['K' + str(x+1)].value)
    KContainer.append(KEntry)
    KContainer[x].grid(row=10, column=x)

#L1 through L12
    LEntry = tk.Entry(window,
                   fg="#4969E1",
                   bg="pink")
    LEntry.insert(0, ws['L' + str(x+1)].value)
    LContainer.append(LEntry)
    LContainer[x].grid(row=11, column=x)

#M1 through M12
    MEntry = tk.Entry(window,
                   fg="#4969E1",
                   bg="pink")
    MEntry.insert(0, ws['M' + str(x+1)].value)
    MContainer.append(MEntry)
    MContainer[x].grid(row=12, column=x)

#N1 through N12
    NEntry = tk.Entry(window,
                   fg="#4969E1",
                   bg="pink")
    NEntry.insert(0, ws['N' + str(x+1)].value)
    NContainer.append(NEntry)
    NContainer[x].grid(row=13, column=x)

#O1 through O12
    OEntry = tk.Entry(window,
                   fg="#4969E1",
                   bg="pink")
    OEntry.insert(0, ws['O' + str(x+1)].value)
    OContainer.append(OEntry)
    OContainer[x].grid(row=14, column=x)

#P1 through P12
    PEntry = tk.Entry(window,
                   fg="#4969E1",
                   bg="pink")
    PEntry.insert(0, ws['P' + str(x+1)].value)
    PContainer.append(PEntry)
    PContainer[x].grid(row=15, column=x)

#Q1 through Q12
    QEntry = tk.Entry(window,
                   fg="#4969E1",
                   bg="pink")
    QEntry.insert(0, ws['Q' + str(x+1)].value)
    QContainer.append(QEntry)
    QContainer[x].grid(row=16, column=x)
#END for loop to initialize values

#Pad left and right for look and feel
CContainer[0].grid(row=2, column=0, padx=(10,0))
DContainer[0].grid(row=3, column=0, padx=(10,0))
EContainer[0].grid(row=4, column=0, padx=(10,0))
FContainer[0].grid(row=5, column=0, padx=(10,0))
GContainer[0].grid(row=6, column=0, padx=(10,0))
HContainer[0].grid(row=7, column=0, padx=(10,0))
IContainer[0].grid(row=8, column=0, padx=(10,0))
JContainer[0].grid(row=9, column=0, padx=(10,0))
KContainer[0].grid(row=10, column=0, padx=(10,0))
LContainer[0].grid(row=11, column=0, padx=(10,0))
MContainer[0].grid(row=12, column=0, padx=(10,0))
NContainer[0].grid(row=13, column=0, padx=(10,0))
OContainer[0].grid(row=14, column=0, padx=(10,0))
PContainer[0].grid(row=15, column=0, padx=(10,0))
QContainer[0].grid(row=16, column=0, padx=(10,0))

CContainer[11].grid(row=2, column=11, padx=(0,10))
DContainer[11].grid(row=3, column=11, padx=(0,10))
EContainer[11].grid(row=4, column=11, padx=(0,10))
FContainer[11].grid(row=5, column=11, padx=(0,10))
GContainer[11].grid(row=6, column=11, padx=(0,10))
HContainer[11].grid(row=7, column=11, padx=(0,10))
IContainer[11].grid(row=8, column=11, padx=(0,10))
JContainer[11].grid(row=9, column=11, padx=(0,10))
KContainer[11].grid(row=10, column=11, padx=(0,10))
LContainer[11].grid(row=11, column=11, padx=(0,10))
MContainer[11].grid(row=12, column=11, padx=(0,10))
NContainer[11].grid(row=13, column=11, padx=(0,10))
OContainer[11].grid(row=14, column=11, padx=(0,10))
PContainer[11].grid(row=15, column=11, padx=(0,10))
QContainer[11].grid(row=16, column=11, padx=(0,10))


#Light Mode Button
LightModeButton = Button(window, text="Light Mode",
                   command=lightMode,
                   fg="#4969E1",
                   bg="pink")
LightModeButton.grid(row=20, column=3)

#Dark Mode Button
DarkModeButton = Button(window, text="Dark Mode",
                   command=darkMode,
                   fg="#4969E1",
                   bg="pink")
DarkModeButton.grid(row=20, column=4)

#Save Button
SaveButton = Button(window, text="Save to File",
                   command=saveFile,
                   fg="#4969E1",
                   bg="pink")
SaveButton.grid(row=20, column=6, pady=5)

#Search Label, Entry, and Button

#Search Label
searchLabel = tk.Label(window,
                   text= "Search Field: ",
                   justify="center",
                   fg="#4969E1",
                   bg="pink")
searchLabel.grid(row=21, column=3, pady=10)

#Search Entry
searchEntry = tk.Entry(window,
                   fg="#4969E1",
                   bg="pink")
searchEntry.grid(row=21, column=4)

#Search Button
searchButton = Button(window, text="Search File",
                   command=searchFile,
                   fg="#4969E1",
                   bg="pink")
searchButton.grid(row=21, column=6)


#Mainloop for display
window.mainloop()

#END main.py